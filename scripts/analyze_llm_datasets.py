"""Analyze LLM Citation Screening datasets for ASReview validation potential."""
import msoffcrypto
import openpyxl
import io
from pathlib import Path
from collections import Counter

DATA_DIR = Path("vendor/llm-citation-screening/Data")
PASSWORD = "oami2025"
OUTPUT_FILE = Path("scripts/llm_dataset_analysis.txt")

def decrypt_xlsx(filepath):
    """Decrypt password-protected xlsx file."""
    with open(filepath, 'rb') as f:
        decrypted = io.BytesIO()
        m = msoffcrypto.OfficeFile(f)
        m.load_key(password=PASSWORD)
        m.decrypt(decrypted)
    return openpyxl.load_workbook(decrypted)

def analyze_file(filepath, out):
    """Analyze a single xlsx file."""
    out.write(f"\n{'='*60}\n")
    out.write(f"FILE: {filepath.name}\n")
    out.write('='*60 + "\n")
    
    wb = decrypt_xlsx(filepath)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        out.write(f"\nSheet: {sheet_name}\n")
        out.write(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} cols\n")
        
        # Get headers (row 1)
        headers = [cell.value for cell in sheet[1]]
        out.write(f"\nColumns:\n")
        for i, h in enumerate(headers):
            if h:
                out.write(f"  [{i}] {h}\n")
        
        # Look for label-like columns
        label_cols = []
        for i, h in enumerate(headers):
            if h and any(x in str(h).lower() for x in ['label', 'include', 'exclude', 'decision', 'incl_', 'rayyan']):
                label_cols.append((i, h))
        
        if label_cols:
            out.write(f"\n** Potential label columns: {label_cols}\n")
            for col_idx, col_name in label_cols:
                values = []
                for row_idx in range(2, sheet.max_row + 1):
                    val = sheet.cell(row=row_idx, column=col_idx+1).value
                    values.append(val)
                counter = Counter(values)
                out.write(f"  {col_name}: {dict(counter.most_common(10))}\n")
        
        # Sample data (rows 2-4)
        out.write("\nSample data (rows 2-4):\n")
        for row_idx in range(2, min(5, sheet.max_row + 1)):
            out.write(f"  Row {row_idx}:\n")
            for col_idx, cell in enumerate(sheet[row_idx]):
                if headers[col_idx] and cell.value:
                    val = str(cell.value)[:80]
                    if len(str(cell.value)) > 80:
                        val += "..."
                    out.write(f"    {headers[col_idx]}: {val}\n")

def main():
    files = sorted(DATA_DIR.glob("*.xlsx"))
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as out:
        out.write(f"LLM Citation Screening Dataset Analysis\n")
        out.write(f"Found {len(files)} files in {DATA_DIR}\n")
        
        for f in files:
            try:
                analyze_file(f, out)
            except Exception as e:
                out.write(f"Error analyzing {f.name}: {e}\n")
    
    print(f"Analysis saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
