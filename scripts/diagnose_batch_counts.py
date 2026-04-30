"""
LLM 一括判定の件数不整合を切り分けるための診断スクリプト。

ユーザー報告の問題:
  - 「未判定: 11214」「一括実行完了 (7732件処理)」「Include: 9411 / Exclude: 9635 (合計 19046)」
    という3つの数値の関係が説明できない。
  - リトライ後も「失敗: 1」が残る。

このスクリプトは以下を集計し、上記の数値が何を表しているかを特定する。

事前準備（ユーザー側）:
  対象スプレッドシートを Google Sheets から
    File > ダウンロード > Microsoft Excel (.xlsx)
  でダウンロードし、debug/TiAb_Review_All_unlabeled.xlsx として保存する。
  （または --xlsx オプションで任意のパスを指定可能）

実行:
  python scripts/diagnose_batch_counts.py
  python scripts/diagnose_batch_counts.py --xlsx path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl  # type: ignore

# Windows コンソールでの日本語出力対策
try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

USER_EMAIL = "koizumi.shiho.4h@kyoto-u.ac.jp"

DEFAULT_XLSX = (
    Path(__file__).resolve().parent.parent / "debug" / "TiAb_Review_All_unlabeled.xlsx"
)


def load_xlsx_sheets(xlsx_path: Path) -> dict[str, list[dict[str, str]]]:
    """xlsx の全シートを {sheet_name: [row_dict, ...]} で返す"""
    if not xlsx_path.exists():
        print(f"[ERROR] {xlsx_path} が見つかりません。", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: dict[str, list[dict[str, str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            out[name] = []
            continue
        headers = [(h or "").strip() if isinstance(h, str) else (str(h) if h is not None else "") for h in header_row]

        sheet_rows: list[dict[str, str]] = []
        for row in rows_iter:
            # 全セルが None の行はスキップ
            if row is None or all(v is None for v in row):
                continue
            d: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = row[i] if i < len(row) else None
                if v is None:
                    d[h] = ""
                elif isinstance(v, str):
                    d[h] = v
                else:
                    d[h] = str(v)
            sheet_rows.append(d)
        out[name] = sheet_rows
    wb.close()
    return out


def section(title: str) -> None:
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def summarize_references(refs: list[dict[str, str]]) -> set[str]:
    section("[1] References タブ")
    total = len(refs)
    ref_ids = [r.get("ref_id", "").strip() for r in refs]
    non_empty = [rid for rid in ref_ids if rid]
    unique_ids = set(non_empty)
    counter = Counter(non_empty)
    duplicates = {rid: c for rid, c in counter.items() if c > 1}

    print(f"総行数:               {total}")
    print(f"ref_id 空でない行数:  {len(non_empty)}")
    print(f"ユニーク ref_id 数:   {len(unique_ids)}")
    print(f"重複している ref_id:  {len(duplicates)}")
    if duplicates:
        sample = list(duplicates.items())[:5]
        print(f"  サンプル(最大5件):  {sample}")

    return unique_ids


def summarize_decisions(decisions: list[dict[str, str]], all_ref_ids: set[str]) -> None:
    section("[2] Decisions タブ 全体")
    total = len(decisions)
    print(f"総行数: {total}")

    by_reviewer_type: dict[str, int] = Counter()
    for d in decisions:
        reviewer_id = d.get("reviewer_id", "").strip()
        if not reviewer_id:
            by_reviewer_type["(empty)"] += 1
        elif reviewer_id.startswith("llm:"):
            by_reviewer_type["llm:*"] += 1
        else:
            by_reviewer_type["human"] += 1
    print(f"reviewer_id 種類別行数: {dict(by_reviewer_type)}")

    # 参照されている ref_id が References に存在するか
    referenced_ids = {d.get("ref_id", "").strip() for d in decisions if d.get("ref_id")}
    missing = referenced_ids - all_ref_ids
    print(f"Decisions に登場し References に存在しない ref_id 数: {len(missing)}")
    if missing:
        print(f"  サンプル: {list(missing)[:5]}")


def summarize_llm_per_execution(decisions: list[dict[str, str]]) -> dict[str, dict]:
    """各 llm:... reviewer_id ごとに行数とユニーク ref_id を集計"""
    section("[3] LLM execution_id (= reviewer_id) ごとの内訳")

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for d in decisions:
        rid = d.get("reviewer_id", "").strip()
        if rid.startswith("llm:"):
            grouped[rid].append(d)

    if not grouped:
        print("  llm:... の reviewer_id を持つ行が見つかりません。")
        return {}

    summary: dict[str, dict] = {}

    print(f"{'execution_id':<70} {'rows':>6} {'uniq_ref':>9} {'dup_pairs':>10} {'incl_pred':>10} {'excl_pred':>10}")
    print("-" * 120)
    for exec_id, rows in sorted(grouped.items()):
        ref_ids = [r.get("ref_id", "").strip() for r in rows]
        uniq = len(set(ref_ids))
        # 同じ (exec_id, ref_id) が複数行になっている数（=ペア重複）
        pair_counter = Counter(ref_ids)
        dup_pairs = sum(c - 1 for c in pair_counter.values() if c > 1)

        # decision 列の include/exclude/pending 内訳
        decisions_breakdown: Counter[str] = Counter()
        # note 内 include_probability の数も数える
        include_prob_count = 0
        for r in rows:
            decisions_breakdown[r.get("decision", "").strip() or "(empty)"] += 1
            note = r.get("note", "")
            if note:
                try:
                    nd = json.loads(note)
                    if isinstance(nd, dict) and "include_probability" in nd:
                        include_prob_count += 1
                except Exception:
                    pass

        print(
            f"{exec_id[:68]:<70} {len(rows):>6} {uniq:>9} {dup_pairs:>10} "
            f"{decisions_breakdown.get('include', 0):>10} {decisions_breakdown.get('exclude', 0):>10}"
        )
        summary[exec_id] = {
            "rows": len(rows),
            "unique_refs": uniq,
            "dup_pairs": dup_pairs,
            "decisions_breakdown": dict(decisions_breakdown),
            "include_prob_count": include_prob_count,
        }

    print()
    print("解説:")
    print("  rows       : Decisions シート上の当該 execution_id の行数")
    print("  uniq_ref   : 上記の中でユニークな ref_id 数")
    print("  dup_pairs  : 同一 (execution_id, ref_id) が二重以上ある場合の超過行数 (rows - uniq_ref)")
    print("  incl_pred  : decision='include' の行数 (閾値確定後のみ非0)")
    print("  excl_pred  : decision='exclude' の行数 (閾値確定後のみ非0)")

    return summary


def summarize_executions(executions: list[dict[str, str]]) -> None:
    section("[4] LLM_Executions タブ")
    if not executions:
        print("  LLM_Executions が空です。")
        return
    print(f"総実行数: {len(executions)}")
    print()
    cols = [
        "execution_id", "execution_type", "timestamp", "status", "is_active",
        "target_count", "include_count", "exclude_count", "include_threshold",
    ]
    # ヘッダ
    widths = {c: max(len(c), 12) for c in cols}
    widths["execution_id"] = 70
    widths["timestamp"] = 22
    print(" ".join(c.ljust(widths[c]) for c in cols))
    print("-" * sum(widths.values()))
    for ex in executions:
        line = []
        for c in cols:
            v = (ex.get(c, "") or "").strip()
            line.append(v[:widths[c]].ljust(widths[c]))
        print(" ".join(line))


def summarize_unjudged_in_any_execution(
    decisions: list[dict[str, str]], all_ref_ids: set[str]
) -> None:
    """全 execution を通じて一度も llm: 判定が付かなかった ref_id を特定（=問題① の永続失敗候補）"""
    section("[5b] 全 execution を通じて一度も判定が付いていない ref_id (=永続失敗候補)")
    judged_ref_ids = {
        d.get("ref_id", "").strip()
        for d in decisions
        if d.get("reviewer_id", "").strip().startswith("llm:")
    }
    judged_ref_ids.discard("")
    never_judged = all_ref_ids - judged_ref_ids
    print(f"全 References:                    {len(all_ref_ids)}")
    print(f"いずれかの llm execution で判定: {len(judged_ref_ids)}")
    print(f"一度も判定が付いていない ref_id: {len(never_judged)}")
    if never_judged:
        for rid in sorted(never_judged):
            print(f"  - {rid}")


def summarize_persistent_failure_in_last_run(
    decisions: list[dict[str, str]],
    all_ref_ids: set[str],
    refs: list[dict[str, str]],
) -> None:
    """画面で見えていた『失敗1件』候補を特定し、その文献の長さや特徴を表示"""
    section("[5c] 直近の execution で失敗していた可能性のある ref_id とその内容")
    judged_ref_ids = {
        d.get("ref_id", "").strip()
        for d in decisions
        if d.get("reviewer_id", "").strip().startswith("llm:")
    }
    judged_ref_ids.discard("")
    never = all_ref_ids - judged_ref_ids
    if not never:
        print("  全ての ref_id は少なくとも一度は判定されています。")
        print("  画面に出ていた『失敗1件』は最後の handleRetryFailed の中で")
        print("  失敗扱いになっただけで、過去のいずれかの run で判定行が既に")
        print("  spreadsheet に保存されている可能性があります。")
        return
    by_id = {r.get("ref_id", "").strip(): r for r in refs}
    for rid in sorted(never):
        r = by_id.get(rid, {})
        title = (r.get("title", "") or "")
        abstract = (r.get("abstract", "") or "")
        print(f"ref_id: {rid}")
        print(f"  title長:    {len(title)} 文字")
        print(f"  abstract長: {len(abstract)} 文字")
        print(f"  title 先頭: {title[:160]!r}")
        print(f"  abstract 先頭: {abstract[:300]!r}")


def summarize_user_view(decisions: list[dict[str, str]], all_ref_ids: set[str]) -> None:
    section(f"[5] ユーザー視点: {USER_EMAIL}")
    user_decisions = [d for d in decisions if d.get("reviewer_id", "").strip() == USER_EMAIL]
    user_decided_refs = {d.get("ref_id", "").strip() for d in user_decisions}
    user_decided_refs.discard("")

    breakdown = Counter((d.get("decision", "") or "").strip() for d in user_decisions)
    print(f"このメールアドレス名義の Decisions 行数: {len(user_decisions)}")
    print(f"判定済みユニーク ref_id 数:               {len(user_decided_refs)}")
    print(f"decision 内訳:                            {dict(breakdown)}")

    unjudged = all_ref_ids - user_decided_refs
    print(f"このユーザーから見た「未判定」ref_id 数:  {len(unjudged)}")
    print()
    print("UI 上の値との比較:")
    print(f"  画面表示の「未判定: 11214」と上記が一致するか確認してください。")
    print(f"  もし上記が 11214 より大きい場合、UI 表示は state.references.length")
    print(f"  (すなわち全件 or 担当割り当て分) を「未判定」とラベルしているだけで、")
    print(f"  実際の未判定数とは別物だという仮説 (β) が裏付けられます。")


def deduce_summary(execution_summary: dict[str, dict], all_ref_ids: set[str]) -> None:
    section("[6] ユーザー報告値との突き合わせ")
    print("ユーザー報告値:")
    print("  未判定:                        11214")
    print("  一括実行完了 件数:             7732")
    print("  Include + Exclude (preview): 9411 + 9635 = 19046")
    print()

    print("コード上、'Include + Exclude' は閾値プレビュー時に")
    print("  state.currentBatchDecisions の長さを使って算出される。")
    print("  - 通常、初回バッチ + リトライ累積で最大でも対象件数を超えないはず。")
    print("  - 19046 > 11214 ということは、何らかの形で同一 ref が複数回")
    print("    currentBatchDecisions に積まれている、もしくは過去実行の")
    print("    閾値変更時に Decisions シート上の重複行を読み込んでいる可能性が高い。")
    print()

    if not execution_summary:
        print("LLM execution が無いため詳細推定不可。")
        return

    print("各 execution_id の rows / uniq_ref を確認し、")
    print("  rows が uniq_ref を大きく上回る execution があれば、")
    print("  そこで appendDecisions が重複 append している = 仮説 γ。")
    print("  どの execution の合計 rows が 19046 に近いかを見ると、")
    print("  画面で表示されていた currentBatchDecisions の出所が特定できる。")
    print()

    total_rows = sum(s["rows"] for s in execution_summary.values())
    total_uniq = sum(s["unique_refs"] for s in execution_summary.values())
    total_dup = sum(s["dup_pairs"] for s in execution_summary.values())
    print(f"全 execution 合計 rows: {total_rows}")
    print(f"全 execution 合計 uniq_ref: {total_uniq}")
    print(f"全 execution 合計 dup_pairs: {total_dup}")

    # 19046 に最も近い execution / 累積を特定
    nearest = min(
        execution_summary.items(),
        key=lambda kv: abs(kv[1]["rows"] - 19046),
        default=None,
    )
    if nearest:
        ex_id, s = nearest
        print(f"単一 execution で 19046 に最も近いのは: {ex_id}  (rows={s['rows']})")


def main() -> None:
    parser = argparse.ArgumentParser(description="LLM 一括判定の件数不整合 診断")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="入力 xlsx のパス")
    args = parser.parse_args()

    print(f"入力ファイル: {args.xlsx}")
    sheets = load_xlsx_sheets(args.xlsx)
    print(f"シート一覧:   {list(sheets.keys())}")

    if "References" not in sheets:
        print("[ERROR] 'References' シートが見つかりません。", file=sys.stderr)
        sys.exit(1)
    if "Decisions" not in sheets:
        print("[ERROR] 'Decisions' シートが見つかりません。", file=sys.stderr)
        sys.exit(1)

    refs = sheets["References"]
    decisions = sheets["Decisions"]
    executions = sheets.get("LLM_Executions", [])

    all_ref_ids = summarize_references(refs)
    summarize_decisions(decisions, all_ref_ids)
    exec_summary = summarize_llm_per_execution(decisions)
    summarize_executions(executions)
    summarize_user_view(decisions, all_ref_ids)
    summarize_unjudged_in_any_execution(decisions, all_ref_ids)
    summarize_persistent_failure_in_last_run(decisions, all_ref_ids, refs)
    deduce_summary(exec_summary, all_ref_ids)


if __name__ == "__main__":
    main()
