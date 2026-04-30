"""
14 個の execution_id に分散している 27024 件の Decisions を、
1 つの canonical execution_id に集約した 11214 行に作り直す。

入力:
  debug/TiAb_Review_All_unlabeled.xlsx

出力 (debug/ 配下):
  consolidated_decisions.csv     差し替え後の Decisions シート全体 (header + 11214 行)
  consolidated_executions.csv    LLM_Executions シート全体 (header + 1 行)
  consolidate_summary.txt        変更内容のサマリ
  consolidate_log.json           ref_id ごとに「どの run のどの判定を採用したか」のログ

ルール:
  - 11213 ref はいずれかの run で判定済み → 最新タイムスタンプの run の判定を採用
  - 残り 1 ref (rayyan-499353097) は GEMINI_API_KEY があれば API で判定、無ければプレースホルダ
  - reviewer_id 列と note JSON の execution_id を canonical に書き換え
  - 他のフィールド (model, reasons, evidence, usageMetadata 等) は元の run の値をそのまま保持
  - decided_at は元の値を保持 (履歴追跡のため)

使い方:
  python scripts/consolidate_decisions.py
"""

from __future__ import annotations

import csv
import json
import os
import sys
import time
import urllib.request
import urllib.error
from collections import defaultdict
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parent.parent
INPUT_XLSX = ROOT / "debug" / "TiAb_Review_All_unlabeled.xlsx"
OUT_DIR = ROOT / "debug"

CANONICAL_EXEC_ID = "llm:gemini-2.5-flash-lite@2026-04-28T05:46:23.972Z"
CANONICAL_MODEL = "gemini-2.5-flash-lite"
MISSING_REF_ID = "rayyan-499353097"

DECISIONS_HEADERS = [
    "decision_id", "ref_id", "reviewer_id", "decision", "reason",
    "labels", "note", "decided_at", "client_version", "source_url",
]
LLM_EXECUTIONS_HEADERS = [
    "execution_id", "execution_type", "timestamp", "model",
    "temperature", "topP", "thinkingLevel",
    "criteria_snapshot", "screening_prompt", "include_threshold",
    "target_count", "include_count", "exclude_count",
    "status", "is_active",
]

# --- Gemini call (rayyan-499353097 用) -----------------------------------

GEMINI_API_BASE = "https://generativelanguage.googleapis.com/v1beta/models"
SCREENING_OUTPUT_SCHEMA = {
    "type": "object",
    "properties": {
        "include_probability": {"type": "number"},
        "reasons": {"type": "array", "items": {"type": "string"}},
        "evidence": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "field": {"type": "string", "enum": ["title", "abstract"]},
                    "quote": {"type": "string"},
                    "start_char": {"type": "integer"},
                    "end_char": {"type": "integer"},
                },
                "required": ["field", "quote", "start_char", "end_char"],
            },
        },
    },
    "required": ["include_probability", "reasons", "evidence"],
}


def load_dotenv() -> None:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k, v)


def call_gemini(model: str, api_key: str, prompt: str, timeout: int = 300) -> dict:
    url = f"{GEMINI_API_BASE}/{model}:streamGenerateContent?key={api_key}"
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
            "responseSchema": SCREENING_OUTPUT_SCHEMA,
        },
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8")

    responses = json.loads(raw)
    full_text = ""
    usage = {}
    for res in responses:
        for cand in res.get("candidates") or []:
            for part in (cand.get("content") or {}).get("parts") or []:
                if part.get("thought") is True:
                    continue
                if part.get("text"):
                    full_text += part["text"]
        if res.get("usageMetadata"):
            usage = res["usageMetadata"]
    parsed = json.loads(full_text)
    return {"parsed": parsed, "usage": usage}


# --- xlsx 読み込み -------------------------------------------------------

def load_xlsx() -> dict:
    print(f"loading {INPUT_XLSX} ...")
    wb = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)

    # References: ref_id, title, abstract を取り出す
    ws = wb["References"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    idx = {name: header.index(name) for name in header if name}
    refs: dict[str, dict[str, str]] = {}
    for r in rows[1:]:
        if not r or not r[idx["ref_id"]]:
            continue
        rid = str(r[idx["ref_id"]]).strip()
        if not rid:
            continue
        refs[rid] = {
            "title": str(r[idx["title"]] or "") if "title" in idx else "",
            "abstract": str(r[idx["abstract"]] or "") if "abstract" in idx else "",
        }

    # Decisions: 全行
    ws = wb["Decisions"]
    rows = list(ws.iter_rows(values_only=True))
    d_header = list(rows[0])
    d_idx = {name: d_header.index(name) for name in d_header if name}
    decisions: list[dict] = []
    for r in rows[1:]:
        if not r:
            continue
        rec = {}
        for k, i in d_idx.items():
            v = r[i]
            rec[k] = "" if v is None else str(v)
        if not rec.get("decision_id"):
            continue
        decisions.append(rec)

    # LLM_Executions: 全行 (canonical の screening_prompt を取り出すため)
    ws = wb["LLM_Executions"]
    rows = list(ws.iter_rows(values_only=True))
    e_header = list(rows[0])
    e_idx = {name: e_header.index(name) for name in e_header if name}
    executions: list[dict] = []
    for r in rows[1:]:
        if not r:
            continue
        rec = {}
        for k, i in e_idx.items():
            v = r[i]
            rec[k] = "" if v is None else str(v)
        if rec.get("execution_id"):
            executions.append(rec)

    print(f"  references:  {len(refs)}")
    print(f"  decisions:   {len(decisions)}")
    print(f"  executions:  {len(executions)}")
    return {"refs": refs, "decisions": decisions, "executions": executions}


# --- 集約ロジック --------------------------------------------------------

def consolidate(decisions: list[dict]) -> tuple[dict[str, dict], dict[str, str]]:
    """ref_id ごとに最新の判定を選び、(canonical_row, source_run_id) を返す"""
    by_ref: dict[str, list[dict]] = defaultdict(list)
    for d in decisions:
        rid = d.get("ref_id", "").strip()
        if not rid:
            continue
        by_ref[rid].append(d)

    canonical: dict[str, dict] = {}
    source: dict[str, str] = {}
    for rid, rows in by_ref.items():
        rows.sort(key=lambda x: x.get("decided_at", ""))
        latest = rows[-1]
        canonical[rid] = latest
        source[rid] = latest.get("reviewer_id", "")
    return canonical, source


def summarize_reasons(reasons: list) -> str:
    """src/lib/llm-processor.ts:summarizeReasons の Python 移植"""
    if not reasons:
        return ""
    if len(reasons) == 1:
        return str(reasons[0])
    return "。".join(str(r) for r in reasons[:3])


def rewrite_for_canonical(row: dict) -> dict:
    """reviewer_id 列と note 内 execution_id を CANONICAL に書き換え。
    reason 列は note.reasons から summarizeReasons で埋め直す
    (本来は閾値確定時に applyThresholdToDecisions が埋めるが、
     シート上での可読性向上のため事前に書き込む)。"""
    new_row = dict(row)
    new_row["reviewer_id"] = CANONICAL_EXEC_ID
    note = row.get("note", "")
    if note:
        try:
            note_obj = json.loads(note)
            if isinstance(note_obj, dict) and note_obj.get("type") == "llm":
                note_obj["execution_id"] = CANONICAL_EXEC_ID
                new_row["note"] = json.dumps(note_obj, ensure_ascii=False)
                new_row["reason"] = summarize_reasons(note_obj.get("reasons") or [])
        except Exception:
            pass
    return new_row


# --- rayyan-499353097 を Gemini で判定 ----------------------------------

def build_prompt(title: str, abstract: str, screening_prompt: str) -> str:
    return (
        f"{screening_prompt}\n\n"
        "## 対象文献\n\n"
        f"**タイトル:**\n{title}\n\n"
        f"**抄録:**\n{abstract or '(抄録なし)'}\n\n"
        "## 出力指示\n"
        "- include_probability: 組み入れ基準に合致する確率を0.0〜1.0で出力\n"
        "- reasons: 判断理由を日本語で短文配列で出力\n"
        "- evidence: タイトルまたは抄録から判断根拠となる部分を正確に抜粋（quote）し、"
        "その開始位置（start_char）と終了位置（end_char）を指定\n\n"
        "注意: quote は title または abstract 内の正確な部分文字列でなければなりません。"
    )


def fetch_missing_decision(refs: dict, executions: list[dict], api_key: str) -> dict | None:
    if MISSING_REF_ID not in refs:
        print(f"[WARN] {MISSING_REF_ID} が References にありません")
        return None
    ref = refs[MISSING_REF_ID]
    canonical_exec = next(
        (e for e in executions if e.get("execution_id") == CANONICAL_EXEC_ID),
        None,
    )
    if canonical_exec is None:
        print(f"[WARN] LLM_Executions に {CANONICAL_EXEC_ID} がありません")
        return None
    screening_prompt = canonical_exec.get("screening_prompt", "").strip()
    if not screening_prompt:
        print("[WARN] screening_prompt が空です")
        return None

    prompt = build_prompt(ref["title"], ref["abstract"], screening_prompt)
    print(f"calling Gemini for {MISSING_REF_ID} ...")
    try:
        result = call_gemini(CANONICAL_MODEL, api_key, prompt)
    except Exception as e:
        print(f"[ERROR] Gemini call failed: {e}")
        return None

    parsed = result["parsed"]
    usage = result.get("usage", {})
    note_obj = {
        "type": "llm",
        "execution_id": CANONICAL_EXEC_ID,
        "model": CANONICAL_MODEL,
        "include_probability": parsed.get("include_probability"),
        "reasons": parsed.get("reasons", []),
        "evidence": parsed.get("evidence", []),
        "prompt_version": "v1.0.0",
        "usageMetadata": usage,
        "_note": "filled by scripts/consolidate_decisions.py to recover from persistent failure",
    }
    decided_at = time.strftime("%Y-%m-%dT%H:%M:%S.000Z", time.gmtime())
    decision_id = _generate_uuid()
    return {
        "decision_id": decision_id,
        "ref_id": MISSING_REF_ID,
        "reviewer_id": CANONICAL_EXEC_ID,
        "decision": "pending",
        "reason": summarize_reasons(parsed.get("reasons") or []),
        "labels": "",
        "note": json.dumps(note_obj, ensure_ascii=False),
        "decided_at": decided_at,
        "client_version": "",
        "source_url": "",
    }


def _generate_uuid() -> str:
    import uuid
    return str(uuid.uuid4())


# --- 出力 ---------------------------------------------------------------

def write_decisions_csv(rows: list[dict], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(DECISIONS_HEADERS)
        for r in rows:
            w.writerow([r.get(c, "") for c in DECISIONS_HEADERS])


def write_executions_csv(executions: list[dict], path: Path) -> None:
    canonical = next(
        (e for e in executions if e.get("execution_id") == CANONICAL_EXEC_ID),
        None,
    )
    if canonical is None:
        print(f"[WARN] canonical execution が無いため LLM_Executions CSV をスキップ")
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(LLM_EXECUTIONS_HEADERS)
        w.writerow([canonical.get(c, "") for c in LLM_EXECUTIONS_HEADERS])


def main() -> None:
    load_dotenv()
    data = load_xlsx()
    refs: dict = data["refs"]
    decisions: list[dict] = data["decisions"]
    executions: list[dict] = data["executions"]

    print()
    print("=== 集約 ===")
    canonical_map, source_map = consolidate(decisions)
    print(f"判定済み ref_id 数:   {len(canonical_map)}")

    rewritten = [rewrite_for_canonical(canonical_map[rid]) for rid in canonical_map]
    rewritten_by_ref = {r["ref_id"]: r for r in rewritten}

    # 欠損 ref を埋める
    missing_refs = set(refs.keys()) - set(canonical_map.keys())
    print(f"欠損 ref_id 数:       {len(missing_refs)}")
    for rid in sorted(missing_refs):
        print(f"  - {rid}")

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    filled_refs: list[str] = []
    placeholder_refs: list[str] = []
    if missing_refs:
        if not api_key:
            print("[WARN] GEMINI_API_KEY 未設定のためプレースホルダ行を出力します")
        for rid in sorted(missing_refs):
            if rid == MISSING_REF_ID and api_key:
                row = fetch_missing_decision(refs, executions, api_key)
                if row:
                    rewritten_by_ref[rid] = row
                    filled_refs.append(rid)
                    continue
            # プレースホルダ: include_probability=0 (exclude 寄り)
            now = time.strftime("%Y-%m-%dT%H:%M:%S.000Z", time.gmtime())
            note_obj = {
                "type": "llm",
                "execution_id": CANONICAL_EXEC_ID,
                "model": CANONICAL_MODEL,
                "include_probability": 0,
                "reasons": ["consolidate_decisions.py のプレースホルダ。手動で見直してください。"],
                "evidence": [],
                "prompt_version": "v1.0.0",
                "_note": "placeholder filled by scripts/consolidate_decisions.py",
            }
            rewritten_by_ref[rid] = {
                "decision_id": _generate_uuid(),
                "ref_id": rid,
                "reviewer_id": CANONICAL_EXEC_ID,
                "decision": "pending",
                "reason": summarize_reasons(note_obj.get("reasons") or []),
                "labels": "",
                "note": json.dumps(note_obj, ensure_ascii=False),
                "decided_at": now,
                "client_version": "",
                "source_url": "",
            }
            placeholder_refs.append(rid)

    # References 順 (= xlsx の出現順) で出力
    final_rows = [rewritten_by_ref[rid] for rid in refs.keys() if rid in rewritten_by_ref]
    print(f"最終出力行数:         {len(final_rows)}")

    # 検証
    seen = {r["ref_id"] for r in final_rows}
    not_in_output = set(refs.keys()) - seen
    if not_in_output:
        print(f"[ERROR] 出力に含まれない ref_id が {len(not_in_output)} 件あります")
        for r in list(not_in_output)[:5]:
            print(f"  - {r}")
        sys.exit(1)

    # 出力
    out_decisions = OUT_DIR / "consolidated_decisions.csv"
    out_executions = OUT_DIR / "consolidated_executions.csv"
    out_summary = OUT_DIR / "consolidate_summary.txt"
    out_log = OUT_DIR / "consolidate_log.json"

    write_decisions_csv(final_rows, out_decisions)
    write_executions_csv(executions, out_executions)

    with out_log.open("w", encoding="utf-8") as f:
        json.dump({
            "canonical_execution_id": CANONICAL_EXEC_ID,
            "ref_to_source_run": source_map,
            "filled_refs_via_gemini": filled_refs,
            "placeholder_refs": placeholder_refs,
        }, f, ensure_ascii=False, indent=2)

    # サマリ
    with out_summary.open("w", encoding="utf-8") as f:
        f.write("=== consolidate_decisions.py サマリ ===\n\n")
        f.write(f"canonical execution_id: {CANONICAL_EXEC_ID}\n\n")
        f.write(f"References 全件:        {len(refs)}\n")
        f.write(f"元の Decisions 行数:    {len(decisions)} (14 execution_ids)\n")
        f.write(f"集約後 Decisions 行数:  {len(final_rows)}\n\n")
        f.write(f"Gemini 再判定で埋めた ref:    {len(filled_refs)} ({filled_refs})\n")
        f.write(f"プレースホルダで埋めた ref:    {len(placeholder_refs)} ({placeholder_refs})\n\n")
        f.write("source run の内訳 (各 ref の判定が来た元 run):\n")
        run_counter: dict[str, int] = defaultdict(int)
        for run in source_map.values():
            run_counter[run] += 1
        for run, n in sorted(run_counter.items(), key=lambda x: -x[1]):
            f.write(f"  {run}: {n}\n")
        f.write("\n")
        f.write(f"出力ファイル:\n")
        f.write(f"  {out_decisions}\n")
        f.write(f"  {out_executions}\n")
        f.write(f"  {out_log}\n")

    print()
    print("=== 完了 ===")
    print(f"  {out_decisions}")
    print(f"  {out_executions}")
    print(f"  {out_summary}")
    print(f"  {out_log}")


if __name__ == "__main__":
    main()
