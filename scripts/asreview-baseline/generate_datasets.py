"""
Generate ASReview validation datasets from LLM Citation Screening data.

This script:
1. Reads CQ data files (title, abstract, etc.) from vendor/llm-citation-screening/Data/
2. Reads Reference_standard to identify included papers for each CQ
3. Matches Reference_standard titles to CQ data using fuzzy matching
4. Outputs labeled JSON datasets for ASReview TypeScript implementation validation
"""
import msoffcrypto
import openpyxl
import io
import json
import unicodedata
import re
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Optional

DATA_DIR = Path("scripts/asreview-baseline/raw_data")
OUTPUT_DIR = Path("scripts/asreview-baseline/datasets")
PASSWORD = "oami2025"

@dataclass
class Reference:
    """A single reference record."""
    id: str
    title: str
    abstract: str
    year: Optional[str]
    journal: Optional[str]
    doi: Optional[str]
    pubmed_id: Optional[str]
    label: int  # 1=include, 0=exclude, -1=unlabeled

def decrypt_xlsx(filepath: Path):
    """Decrypt password-protected xlsx file."""
    with open(filepath, 'rb') as f:
        decrypted = io.BytesIO()
        m = msoffcrypto.OfficeFile(f)
        m.load_key(password=PASSWORD)
        m.decrypt(decrypted)
    return openpyxl.load_workbook(decrypted)

def normalize_title(title: str) -> str:
    """Normalize title for fuzzy matching."""
    if not title:
        return ""
    # Unicode normalize
    title = unicodedata.normalize('NFKC', title)
    # Lowercase
    title = title.lower()
    # Remove punctuation and extra whitespace
    title = re.sub(r'[^\w\s]', ' ', title)
    title = re.sub(r'\s+', ' ', title).strip()
    return title

def load_cq_data(cq_num: int) -> list[Reference]:
    """Load CQ data file and return list of Reference objects."""
    filepath = DATA_DIR / f"CQ{cq_num}_data.xlsx"
    wb = decrypt_xlsx(filepath)
    sheet = wb.active
    
    # Row 2 is headers, data starts at row 3
    headers = [cell.value for cell in sheet[2]]
    
    # Build column index map
    col_map = {h: i for i, h in enumerate(headers) if h}
    
    records = []
    for row_idx in range(3, sheet.max_row + 1):
        row = [cell.value for cell in sheet[row_idx]]
        
        title = row[col_map.get('title', 1)] or ""
        abstract = row[col_map.get('abstract', 15)] or ""
        
        # Skip records without title
        if not title.strip():
            continue
        
        ref = Reference(
            id=row[col_map.get('key', 0)] or f"row-{row_idx}",
            title=str(title).strip(),
            abstract=str(abstract).strip(),
            year=str(row[col_map.get('year', 2)]) if row[col_map.get('year', 2)] else None,
            journal=str(row[col_map.get('journal', 5)]) if row[col_map.get('journal', 5)] else None,
            doi=str(row[col_map.get('doi', 17)]) if row[col_map.get('doi', 17)] else None,
            pubmed_id=str(row[col_map.get('pubmed_id', 19)]) if row[col_map.get('pubmed_id', 19)] else None,
            label=-1  # Will be set later
        )
        records.append(ref)
    
    return records

def load_reference_standard(cq_num: int, stage: str = "1st") -> list[str]:
    """Load Reference_standard titles for a CQ.
    
    Note: Reference_standard_data.xlsx is now a standard unencrypted XLSX.
    """
    filepath = DATA_DIR / "Reference_standard_data.xlsx"
    wb = openpyxl.load_workbook(filepath)
    
    sheet_name = f"CQ{cq_num}_{stage}"
    if sheet_name not in wb.sheetnames:
        print(f"  Warning: Sheet {sheet_name} not found")
        return []
    
    sheet = wb[sheet_name]
    
    # Row 1 is headers (Title, Year, Journal), data starts at row 2
    titles = []
    for row_idx in range(2, sheet.max_row + 1):
        try:
            title = sheet.cell(row=row_idx, column=1).value
            if title:
                titles.append(str(title).strip())
        except Exception:
            continue
    
    return titles

def match_labels(records: list[Reference], included_titles: list[str]) -> tuple[int, int]:
    """Match Reference_standard titles to CQ records and set labels."""
    # Normalize all included titles
    normalized_included = {normalize_title(t): t for t in included_titles}
    
    matched = 0
    unmatched_titles = list(included_titles)
    
    for ref in records:
        norm_title = normalize_title(ref.title)
        
        # Check for exact normalized match
        if norm_title in normalized_included:
            ref.label = 1
            matched += 1
            # Remove from unmatched list
            original = normalized_included[norm_title]
            if original in unmatched_titles:
                unmatched_titles.remove(original)
        else:
            ref.label = 0
    
    # Report unmatched Reference_standard titles
    if unmatched_titles:
        print(f"  Unmatched Reference_standard titles ({len(unmatched_titles)}):")
        for t in unmatched_titles[:5]:
            print(f"    - {t[:80]}...")
        if len(unmatched_titles) > 5:
            print(f"    ... and {len(unmatched_titles) - 5} more")
    
    return matched, len(included_titles)

def generate_dataset(cq_num: int):
    """Generate labeled dataset for a single CQ."""
    print(f"\n=== Processing CQ{cq_num} ===")
    
    # Load CQ data
    records = load_cq_data(cq_num)
    print(f"  Loaded {len(records)} records from CQ{cq_num}_data.xlsx")
    
    # Load Reference_standard (use 1st screen for primary labels)
    included_titles_1st = load_reference_standard(cq_num, "1st")
    included_titles_2nd = load_reference_standard(cq_num, "2nd")
    print(f"  Reference_standard: {len(included_titles_1st)} (1st), {len(included_titles_2nd)} (2nd)")
    
    # Use 1st screening results as labels (larger set, more useful for ML)
    matched, total = match_labels(records, included_titles_1st)
    print(f"  Matched: {matched}/{total} ({100*matched/total:.1f}%)")
    
    # Calculate statistics
    included_count = sum(1 for r in records if r.label == 1)
    excluded_count = sum(1 for r in records if r.label == 0)
    
    # Build output
    output = {
        "dataset": f"CQ{cq_num}",
        "source": "llm-assisted_citation_screening",
        "label_source": f"Reference_standard_data.xlsx (CQ{cq_num}_1st)",
        "statistics": {
            "total": len(records),
            "included": included_count,
            "excluded": excluded_count,
            "prevalence": round(included_count / len(records), 4) if records else 0,
            "reference_standard_count": total,
            "match_rate": round(matched / total, 4) if total else 0
        },
        "records": [asdict(r) for r in records]
    }
    
    # Write output
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"cq{cq_num}_labeled.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"  Output: {output_path}")
    print(f"  Stats: {included_count} included, {excluded_count} excluded, prevalence={output['statistics']['prevalence']:.2%}")
    
    return output

def main():
    print("Generating ASReview validation datasets from LLM Citation Screening data")
    print("=" * 60)
    
    summary = []
    for cq_num in range(1, 6):
        try:
            result = generate_dataset(cq_num)
            summary.append({
                "cq": f"CQ{cq_num}",
                **result["statistics"]
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"  ERROR: {e}")
    
    # Write summary
    summary_path = OUTPUT_DIR / "summary.json"
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    
    print("\n" + "=" * 60)
    print("Summary:")
    print(f"{'CQ':<6} {'Total':<8} {'Incl':<6} {'Excl':<8} {'Prev':<8} {'Match':<8}")
    for s in summary:
        print(f"{s['cq']:<6} {s['total']:<8} {s['included']:<6} {s['excluded']:<8} {s['prevalence']:<8.2%} {s['match_rate']:<8.2%}")
    
    print(f"\nDatasets saved to: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
